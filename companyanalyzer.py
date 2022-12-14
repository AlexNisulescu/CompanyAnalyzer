from datetime import date # Used to get current date
from os import path # Used to check if file exists
import pandas as pd  # Used to store data in a dataframe
import yfinance as yf  # Used for financial data
from openpyxl import load_workbook # Used for importing excel files

# This function is used to return Earnings per share of a company
def earning_per_share_calculator(price_to_earnings, current_price):
    return current_price / price_to_earnings

# This function is used to calculate the growth rate (%) of a company
def growth_rate_calculator(history_price, current_price):
    return ((current_price - history_price) / history_price)

# This function is used to calculate the intrinsic value of a company
def intrinsic_value_calculator(price_to_earnings, current_price, ten_years_price):
    # V = [ EPS * (8.5 + 2 * g) * 4.4 ] / Y , where:
    # EPS = earnings_per_share
    # g = growth_rate
    # Y = bond_yield
    Y = 4.3
    earnings_per_share = earning_per_share_calculator(price_to_earnings, current_price)
    growth_rate = growth_rate_calculator(ten_years_price, current_price)
    return int((earnings_per_share * (8.5 + (2 * growth_rate)) * 4.4) / Y)

# This function is used to calculate the margin of safety of a company
def margin_of_safety(intrinsic_value, current_market_price):
    # MS = ( V - CMP ) / V , where:
    # V = intrinsic_value
    # CMP = current_market_price
    return "{:.2f}".format(float(intrinsic_value - current_market_price) / intrinsic_value * 100)

# This function is used to read the list of tickers
def read_tickers(filename):
    file = open(filename, "r")
    tkrs = file.read().splitlines()
    return tkrs

# This function creates the dataframe structure that will store the data
def create_df():
    df = pd.DataFrame(columns=['Name', 'Sector', 'Industry', 'MarketCap',
                        'P/E', 'Current Price', 'Intrinsic Value', 'Margin of Safety',
                        'Debt to Equity', 'Profit Margins', 'Gross Margins', 
                        'Five Years Average Dividend Yeald', 'Last Dividend Value',
                        'Worth', 'Total Cash', 'Total Assets', 'Total Debt', 
                        'Forward P/E'])
    return df

# This function is used to calculate the worth of a company using cash/assets/debt
def calculate_worth(cash, assets, debt):
    if cash is None:
        cash = 0
    if assets is None:
        assets = 0
    if debt is None:
        debt = 0 
    worth = cash + assets - debt
    return worth

# This function returns a new row for the dataframe when P/E is missing
def get_new_row_without_intrinsic_value(ticker, details, worth):
    new_row = {'Name':ticker, 'Sector':details.info['sector'], 'Industry':
    details.info['industry'], 'MarketCap':details.info['marketCap'],
    'Current Price':details.info['currentPrice'], 'Debt to Equity':
    details.info['debtToEquity'],'Profit Margins':
    details.info['profitMargins'], 'Gross Margins':
    details.info['grossMargins'], 'Five Years Average Dividend Yeald':
    details.info['fiveYearAvgDividendYield'], 'Last Dividend Value':
    details.info['lastDividendValue'], 'Worth':worth ,
    'Total Cash':details.info['totalCash'],'Total Assets':
    details.info['totalAssets'], 'Total Debt':details.info['totalDebt'],
    'Forward P/E':details.info['forwardPE']}
    # Formats the new_row as a dataframe
    new_row = pd.DataFrame([new_row], index=['Name'])
    return new_row

# This function returns a new row for the dataframe
def get_new_row(ticker, details, intrinsic_value, safety_margin, worth):
    new_row = {'Name':ticker, 'Sector':details.info['sector'], 'Industry':
    details.info['industry'], 'MarketCap':details.info['marketCap'],
    'P/E':details.info['trailingPE'], 'Current Price':
    details.info['currentPrice'], 'Intrinsic Value':intrinsic_value,
    'Margin of Safety':safety_margin,'Debt to Equity':details.info['debtToEquity'],
    'Profit Margins':details.info['profitMargins'], 'Gross Margins':
    details.info['grossMargins'], 'Five Years Average Dividend Yeald':
    details.info['fiveYearAvgDividendYield'], 'Last Dividend Value':
    details.info['lastDividendValue'], 'Worth':worth ,
    'Total Cash':details.info['totalCash'],'Total Assets':
    details.info['totalAssets'], 'Total Debt':details.info['totalDebt'],
    'Forward P/E':details.info['forwardPE']}
    # Formats the new_row as a dataframe
    new_row = pd.DataFrame([new_row], index=['Name'])
    return new_row

# This function gets the details of the companies and stores them in a dataframe
def get_details(tkrs):
    df = create_df()
    # Parses the whole list of tickers
    for ticker in tkrs:
        # Requests the company details
        print("Downloading data for " + ticker)
        details = yf.Ticker(ticker)
        try:
            try:
                details.info['totalCash'] = int(details.info['totalCash'])
            except TypeError:
                details.info['totalCash'] = 0
            try:
                details.info['totalAssets'] = int(details.info['totalAssets'])
            except TypeError:
                details.info['totalAssets'] = 0
            try:
                details.info['totalDebt'] = int(details.info['totalDebt'])
            except TypeError:
                details.info['totalDebt'] = 0
            # Calculates a company worth
            worth = calculate_worth(details.info['totalCash'],
                    details.info['totalAssets'], details.info['totalDebt'])
            # Saves the stock price ten years ago
            ten_years_price = details.history(period='10y')
            # Saves current price
            current_price = details.history(period='1d')
            # Calculates the intrinsic value of the company
            # Creates a new row for the dataframe
        except:
            print("Ticker for " + ticker + " doesn't exists, or data could not be downloaded.")
            continue
        try:
            intrinsic_value = intrinsic_value_calculator(details.info['trailingPE'],
                          current_price['Close'][0], ten_years_price['Close'][0])
            # Calculates the safety margin
            safety_margin = margin_of_safety(intrinsic_value, current_price['Close'][0])
            # Formats the safety margin in a more human readable format
            safety_margin = str(safety_margin) + "%"
            safety_margin = safety_margin.replace(".", ",")
        except:
            new_row = get_new_row_without_intrinsic_value(ticker, details, worth)
            df = pd.concat([df, new_row], ignore_index=True)
            print("Intrinsic value could not be compiled.. Missing P/E..")
            print("This usually occurs for the new listed companies..")
            continue
        new_row = get_new_row(ticker, details, intrinsic_value, safety_margin, worth)
        # Appends the new row to the dataframe
        df = pd.concat([df, new_row], ignore_index=True)
    # Sorts the companies by MarketCap
    df.sort_values(by=['MarketCap'], inplace=True)
    # Resets the index
    df = df.reset_index()
    # Drop the index column
    df.drop('index', axis=1, inplace=True)
    return df

# This functions saves the dataframe to an excel
def save_dataframe_to_excell(df):
    today = date.today()
    if (path.exists('files/companies.xlsx')):
        wb = load_workbook('files/companies.xlsx', read_only=False)
        if today.strftime("%d.%m.%Y") not in wb.sheetnames:
            with pd.ExcelWriter('files/companies.xlsx',
                            mode='a') as writer:
                df.to_excel(writer, sheet_name=today.strftime("%d.%m.%Y"))
        else:
            print("Attention! Data already downloaded today...")
            print("Overwriteing it...")
            wb[today.strftime("%d.%m.%Y")].title="old"
            wb.save('files/companies.xlsx')
            with pd.ExcelWriter('files/companies.xlsx',
                        mode='a') as writer:
                df.to_excel(writer, sheet_name=today.strftime("%d.%m.%Y"))
            wb = load_workbook('files/companies.xlsx', read_only=False)
            del wb['old']
            wb.save('files/companies.xlsx')
    else:
        df.to_excel('files/companies.xlsx', sheet_name=today.strftime("%d.%m.%Y"))

# This functions saves the dataframe to a csv
def save_dataframe_to_csv(df):
    df.to_csv('files/companies.csv')

# Reading the tickers
tickers = read_tickers("files/Tickers")
# Getting companies details
companydf = get_details(tickers)
# Starting the index count from 1
companydf.index += 1
# Saving the details to an Excell file
save_dataframe_to_excell(companydf)
save_dataframe_to_csv(companydf)